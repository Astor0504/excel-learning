#!/usr/bin/env python3
"""把兩個 xlsx 抽取成 xlsx-content.js（lesson-slug → 結構化內容）。"""
from openpyxl import load_workbook
import json, re
from pathlib import Path

ROOT = Path(__file__).parent
INTERACTIVE = ROOT / 'Excel互動練習_macOS版.xlsx'
PROFESSIONAL = ROOT / 'Excel專業養成_macOS版.xlsx'
OUT = ROOT / 'xlsx-content.js'


def rows_of(ws):
    return list(ws.iter_rows(values_only=True))


def trim(v):
    if v is None:
        return ''
    if isinstance(v, str):
        return v.strip()
    return v


def is_blank_row(row):
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in row)


# ---------- 互動練習 sheets（固定格式：資料表 + 題組 with hidden answer in col H）----------
def parse_interactive(ws):
    rows = rows_of(ws)
    title = trim(rows[0][0])
    subtitle = trim(rows[1][0])
    # 練習資料：rows[3] 是【練習資料】，rows[4] 是 header，rows[5..10] 是資料（5 row）
    data_header = [trim(c) for c in rows[4][:7]]
    data_rows = []
    for r in rows[5:11]:
        if is_blank_row(r):
            break
        data_rows.append([trim(c) for c in r[:7]])
    # 題組：偵測兩種 header
    # 格式 A: ('#', '難度', '任務描述', '你的答案 ▼', '結果', '提示...', None, None) → num=col0
    # 格式 B: ('難度', '題號', '任務描述', '你的答案 ▼', '結果', '提示...', None, None) → diff=col0, num=col1
    tasks = []
    header_type = None
    for r in rows:
        c0, c1 = trim(r[0]), trim(r[1])
        if c0 == '#' and c1 == '難度':
            header_type = 'A'; continue
        if c0 == '難度' and c1 == '題號':
            header_type = 'B'; continue
        if header_type is None:
            continue
        if header_type == 'A' and isinstance(r[0], (int, float)):
            num, diff, desc, _, _, hint, _, answer = r[:8]
        elif header_type == 'B' and isinstance(r[1], (int, float)):
            diff, num, desc, _, _, hint, _, answer = r[:8]
        else:
            continue
        tasks.append({
            'num': int(num),
            'difficulty': trim(diff),
            'desc': trim(desc),
            'hint': trim(hint),
            'answer': trim(answer),
        })
    return {
        'title': title,
        'subtitle': subtitle,
        'dataHeader': data_header,
        'dataRows': data_rows,
        'tasks': tasks,
    }


# ---------- 專業養成 函式 sheets（資料表 + 題組 with 提示+參考答案+原理解說）----------
def parse_pro_function(ws):
    rows = rows_of(ws)
    title = trim(rows[0][0])
    subtitle = trim(rows[1][0])
    # 找【練習資料】後的 header 與資料
    data_header = []
    data_rows = []
    tasks = []
    state = 'init'
    for i, r in enumerate(rows):
        cell0 = trim(r[0])
        if isinstance(cell0, str) and cell0.startswith('【練習資料】'):
            data_header = [trim(c) for c in rows[i+1][:7]]
            for rr in rows[i+2:i+12]:
                if is_blank_row(rr):
                    break
                if isinstance(rr[0], str) and rr[0].startswith('【'):
                    break
                data_rows.append([trim(c) for c in rr[:7]])
        # 題組從「#, ✓, 任務描述, ⏱, 提示, 你的答案, 參考答案, 原理解說」之後
        if r[0] == '#' and r[1] == '✓':
            for rr in rows[i+1:]:
                if rr[0] is None:
                    continue
                if isinstance(rr[0], (int, float)) or (isinstance(rr[0], str) and re.match(r'^[🟢🔵🟡🟠🔴]\d+$', rr[0])):
                    tasks.append({
                        'numLabel': trim(rr[0]),
                        'desc': trim(rr[2]),
                        'time': trim(rr[3]),
                        'hint': trim(rr[4]),
                        'answer': trim(rr[6]),
                        'explain': trim(rr[7]) if len(rr) > 7 else '',
                    })
    return {
        'title': title,
        'subtitle': subtitle,
        'dataHeader': data_header,
        'dataRows': data_rows,
        'tasks': tasks,
    }


# ---------- 專業養成 知識型 sheets（章節 + key/value 表格 + 動手練習）----------
def parse_pro_knowledge(ws):
    rows = rows_of(ws)
    title = trim(rows[0][0])
    subtitle = trim(rows[1][0])
    sections = []
    hands_tasks = []
    cur = None
    in_hands = False
    for r in rows[2:]:
        cell0 = trim(r[0])
        rest = [trim(c) for c in r[1:]] if len(r) > 1 else []
        rest_empty = all(c == '' for c in rest)
        # 動手練習區塊
        if isinstance(cell0, str) and '動手練習' in cell0:
            in_hands = True; cur = None; continue
        if in_hands:
            if cell0 == '#' and trim(r[1]) == '✓':
                continue
            if isinstance(cell0, (int, float)):
                hands_tasks.append({
                    'num': int(cell0),
                    'difficulty': trim(r[2]),
                    'desc': trim(r[3]) if len(r) > 3 else '',
                })
            continue
        if is_blank_row(r):
            continue
        # section header: col0 有值，其餘都空（且像章節標題）
        if cell0 and rest_empty and isinstance(cell0, str) and (
            cell0.startswith(('📌','🔧','⚡','❌','💡','🎯','📊','🏆','📋','🛡️','✅','📈','🔗','📝','⚠️','🔥','📐','🎨','⌨️','🚀','💼','🏃','📂','🧩','🔬','⚙️'))
            or cell0.endswith('】') or '步驟' in cell0 or '技巧' in cell0 or '錯誤' in cell0 or '原則' in cell0
        ):
            cur = {'title': cell0, 'items': []}
            sections.append(cur)
        elif cur is not None and cell0:
            cur['items'].append([cell0] + [c for c in rest if c != ''])
        elif cell0:
            if not sections or sections[-1]['title'] != '說明':
                cur = {'title': '說明', 'items': []}
                sections.append(cur)
            sections[-1]['items'].append([cell0] + [c for c in rest if c != ''])
    return {'title': title, 'subtitle': subtitle, 'sections': sections, 'handsTasks': hands_tasks}


# ---------- 快捷鍵大全 ----------
def parse_shortcuts(ws):
    rows = rows_of(ws)
    title = trim(rows[0][0])
    subtitle = trim(rows[1][0])
    intro = trim(rows[2][0])
    groups = []
    cur = None
    for r in rows[3:]:
        cell0 = trim(r[0])
        if is_blank_row(r):
            continue
        # 群組標題（單欄文字，含 emoji 開頭）
        if cell0 and all(trim(c) == '' for c in r[1:]):
            # 略過動手練習區塊：放在快捷鍵頁但內容只有打勾欄+難度，沒有實際描述
            if isinstance(cell0, str) and '動手練習' in cell0:
                cur = None
                continue
            cur = {'title': cell0, 'items': []}
            groups.append(cur)
        elif cell0 == 'macOS 快捷鍵' or cell0 == '快捷鍵':
            continue  # header row, skip
        elif cur is not None and cell0:
            cur['items'].append({
                'key': cell0,
                'desc': trim(r[1]) if len(r) > 1 else '',
                'note': trim(r[2]) if len(r) > 2 else '',
            })
    return {'title': title, 'subtitle': subtitle, 'intro': intro, 'groups': groups}


# ---------- VBA sheets ----------
def parse_vba(ws):
    rows = rows_of(ws)
    title = trim(rows[0][0])
    subtitle = trim(rows[1][0])
    warning = trim(rows[2][0])
    quick_ops = []
    micro_tasks = []
    practice_tasks = []
    state = 'header'
    cur_task = None

    for i, r in enumerate(rows[3:], start=3):
        cell0 = trim(r[0])
        time_marker = trim(r[3]) if len(r) > 3 else ''
        # 快速操作指南
        if cell0 == '⌨️ 快速操作指南':
            state = 'quickops'; continue
        if state == 'quickops' and cell0 and trim(r[1]):
            quick_ops.append({'op': cell0, 'how': trim(r[1])}); continue
        # 微任務開始：col0 有「微任務」+ col3 有「⏱」
        if isinstance(cell0, str) and (cell0.startswith('微任務') or cell0.startswith('實戰') or cell0.startswith('進階')) and isinstance(time_marker, str) and '⏱' in time_marker:
            if cur_task: micro_tasks.append(cur_task)
            cur_task = {'title': cell0, 'time': time_marker, 'code': [], 'tip': ''}
            state = 'micro'; continue
        if state == 'micro' and cur_task is not None:
            if isinstance(cell0, str) and cell0.startswith('💡'):
                cur_task['tip'] = cell0
            elif cell0 != '' and not is_blank_row(r):
                cur_task['code'].append(cell0)
            elif is_blank_row(r):
                pass
        # 動手練習表
        if isinstance(cell0, str) and '動手練習' in cell0:
            if cur_task: micro_tasks.append(cur_task); cur_task = None
            state = 'practice'; continue
        if state == 'practice':
            if r[0] == '#' and r[1] == '✓':
                continue
            if isinstance(r[0], (int, float)):
                practice_tasks.append({
                    'num': int(r[0]),
                    'difficulty': trim(r[2]),
                    'desc': trim(r[3]),
                })
    if cur_task and cur_task not in micro_tasks:
        micro_tasks.append(cur_task)

    return {
        'title': title, 'subtitle': subtitle, 'warning': warning,
        'quickOps': quick_ops, 'microTasks': micro_tasks, 'practiceTasks': practice_tasks,
    }


# ---------- 儀表板 ----------
def parse_dashboard(ws):
    rows = rows_of(ws)
    stages = {}
    badges = []
    strategies = []
    start_point = []
    cur_phase = None
    state = None
    for r in rows:
        c0 = trim(r[0])
        if isinstance(c0, str):
            if '成就徽章' in c0:
                state = 'badges'; continue
            if '學習策略' in c0:
                state = 'strategies'; continue
            if '你的起點' in c0:
                state = 'start'; continue
            if (c0.startswith('🏃') or c0.startswith('💼') or c0.startswith('📐')
                or c0.startswith('⚡') or c0.startswith('🚀') or c0.startswith('🔧')) and 'Phase' in c0:
                cur_phase = c0; state = 'stages'; continue
        if state == 'stages' and isinstance(c0, str) and c0.startswith('第') and trim(r[1]):
            stages[trim(r[1])] = {
                'phase': cur_phase, 'stage': c0,
                'topics': trim(r[2]), 'difficulty': trim(r[3]),
                'taskCount': trim(r[4]), 'time': trim(r[5]), 'xp': trim(r[6]),
            }
        elif state == 'badges':
            if c0 == '徽章' or c0 == '' or c0 is None: continue
            if isinstance(c0, str) and ('解鎖' in c0 or 'XP' in c0 or c0 == '狀態'):
                continue
            cond = trim(r[1]); xp = trim(r[2])
            if cond and xp:
                badges.append({'name': c0, 'condition': cond, 'xp': xp})
        elif state == 'strategies':
            if c0 and isinstance(c0, str):
                desc = trim(r[2]) if len(r) > 2 else ''
                if desc:
                    strategies.append({'title': c0, 'desc': desc})
        elif state == 'start':
            if c0 and isinstance(c0, str):
                desc = trim(r[2]) if len(r) > 2 else ''
                if desc:
                    start_point.append({'stage': c0, 'desc': desc})
    return {'stages': stages, 'badges': badges, 'strategies': strategies, 'startPoint': start_point}


# ---------- Main ----------
inter_wb = load_workbook(INTERACTIVE, data_only=False)
pro_wb = load_workbook(PROFESSIONAL, data_only=False)

interactive = {sn: parse_interactive(inter_wb[sn]) for sn in inter_wb.sheetnames}

# 專業養成 sheet 分類
pro_function_sheets = ['📊 基礎函式','📐 條件判斷','📈 條件統計','🔍 查找比對','📝 文字日期','📊 陣列動態','🧮 進階函式','🏆 綜合挑戰']
pro_knowledge_sheets = ['📊 樞紐分析表','🎨 條件式格式化','✅ 資料驗證','📈 圖表設計','🔗 命名範圍&表格','⚡ Power Query','🛡️ 保護與安全','🧩 假設分析']
pro_vba_sheets = ['⚙️ VBA基礎','🔧 VBA進階','🏗️ VBA實戰']

professional = {}
for sn in pro_function_sheets:
    professional[sn] = parse_pro_function(pro_wb[sn])
for sn in pro_knowledge_sheets:
    professional[sn] = parse_pro_knowledge(pro_wb[sn])
for sn in pro_vba_sheets:
    professional[sn] = parse_vba(pro_wb[sn])
professional['⌨️ 快捷鍵大全'] = parse_shortcuts(pro_wb['⌨️ 快捷鍵大全'])

dashboard_full = parse_dashboard(pro_wb['🎯 儀表板'])
dashboard = dashboard_full['stages']  # 向後相容

# Lesson slug → 對應 sheets
LESSON_MAP = {
    'P1-01': {'shortcuts': '⌨️ 快捷鍵大全'},
    'P1-02': {'pro': '📊 基礎函式', 'inter': '📊 基礎函式'},
    'P1-03': {'pro': '📐 條件判斷', 'inter': '📐 條件判斷'},
    'P2-01': {'pro': '📈 條件統計', 'inter': '📈 條件統計'},
    'P2-02': {'pro': '🔍 查找比對', 'inter': '🔍 查找比對'},
    'P2-03': {'knowledge': '📊 樞紐分析表'},
    'P2-04': {'knowledge': '🎨 條件式格式化'},
    'P3-01': {'knowledge': '✅ 資料驗證'},
    'P3-02': {'pro': '📝 文字日期', 'inter': '✂️ 文字與日期'},
    'P3-03': {'knowledge': '📈 圖表設計'},
    'P3-04': {'knowledge': '🔗 命名範圍&表格'},
    'P3-05': {'knowledge': '🛡️ 保護與安全'},
    'P4-01': {'pro': '📊 陣列動態', 'inter': '🔢 陣列與動態'},
    'P4-02': {'pro': '🧮 進階函式', 'inter': '🏆 進階綜合'},
    'P4-03': {'knowledge': '🧩 假設分析'},
    'P4-04': {'knowledge': '⚡ Power Query'},
    'P5-01': {'vba': '⚙️ VBA基礎'},
    'P5-02': {'vba': '🔧 VBA進階'},
    'P5-03': {'vba': '🏗️ VBA實戰'},
    'P5-04': {'pro': '🏆 綜合挑戰'},
}

lessons = {}
for slug, refs in LESSON_MAP.items():
    obj = {}
    if 'shortcuts' in refs:
        obj['shortcuts'] = professional[refs['shortcuts']]
    if 'pro' in refs:
        obj['pro'] = professional[refs['pro']]
        # 找儀表板對應
        sheet_name = refs['pro']
        if sheet_name in dashboard:
            obj['meta'] = dashboard[sheet_name]
    if 'inter' in refs:
        obj['inter'] = interactive[refs['inter']]
    if 'knowledge' in refs:
        obj['knowledge'] = professional[refs['knowledge']]
        sn = refs['knowledge']
        if sn in dashboard:
            obj['meta'] = dashboard[sn]
    if 'vba' in refs:
        obj['vba'] = professional[refs['vba']]
        sn = refs['vba']
        if sn in dashboard:
            obj['meta'] = dashboard[sn]
    lessons[slug] = obj

# 寫入 JS 檔
out_data = {
    'lessons': lessons,
    'dashboard': dashboard,
    'badges': dashboard_full['badges'],
    'strategies': dashboard_full['strategies'],
    'startPoint': dashboard_full['startPoint'],
}
js = '// Auto-generated from xlsx files. Do not edit by hand.\n'
js += 'window.XLSX_CONTENT = ' + json.dumps(out_data, ensure_ascii=False, indent=1) + ';\n'
OUT.write_text(js, encoding='utf-8')
print(f'✅ Wrote {OUT} ({len(js):,} bytes, {len(lessons)} lessons)')
